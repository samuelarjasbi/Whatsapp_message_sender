import urllib.parse
import time
import sys
from PyQt5 import QtGui
import qtmodern.windows
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QLabel,
    QLineEdit,
    QPushButton,
    QVBoxLayout,
    QFileDialog,
    QMessageBox,
    QListWidget,
    QListWidgetItem,
    QWidget,
)
from PyQt5.QtCore import QThread, pyqtSignal, QObject, Qt
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QLabel,
    QLineEdit,
    QPushButton,
    QVBoxLayout,
    QFileDialog,
    QMessageBox,
    QListWidget,
    QListWidgetItem,
    QWidget,
)



dark_stylesheet = """
QMainWindow {
    background-color: #121212;
}

QMainWindow::title {
    background-color: #121212;
    color: #FFFFFF;
    border: none;
    padding: 6px;
}

QLabel {
    font-family: Roboto, sans-serif;  /* Specify the desired font name or use system fonts */
    font-size: 12px;  /* Adjust the font size as needed */
    font-weight: bold;
    color: #FFFFFF;
}

QLineEdit {
    background-color: #2e2e2e;
    color: #FFFFFF;
    border: none;
}

QPushButton {
    font-family: Roboto, sans-serif;  /* Specify the desired font name or use system fonts */
    font-size: 12px;  /* Adjust the font size as needed */
    font-weight: light;
    background-color: #332940;
    color: #FFFFFF;
    border: none;
    padding: 6px 12px;
    border-radius: 3px;
}

QPushButton:hover {
    background-color: #534a5d;
}

QPushButton:pressed {
    background-color: #444444;
}

QLineEdit::placeholder { 
    color: #888888;
    font-family: Roboto, sans-serif;
    font-weight: bold;
    
}

QListWidget {
    background-color: #222222;
    color: #FFFFFF;
    border: none;
}
"""


class MessageSendingThread(QThread):
    update_list = pyqtSignal(str)

    def __init__(self, driver, wait, recipients, message):
        super().__init__()
        self.driver = driver
        self.wait = wait
        self.recipients = recipients
        self.message = message

    def run(self):
        for recipient in self.recipients:
            send_message(self.driver, self.wait, recipient, self.message)
            self.update_list.emit(recipient)

        self.driver.close()


def initialize_driver():
    options = Options()
    options.add_argument("--ignore-local-proxy")
    return webdriver.Chrome(options=options)


def send_message(driver, wait, recipient, message):
    wait = WebDriverWait(driver, 10) 
    phone = urllib.parse.quote(recipient)

    if isinstance(recipient, float):
        int(recipient)
        
    msg = urllib.parse.quote(message)
    url = (
        "https://web.whatsapp.com/send?phone="
        + phone[:-2]
        + "&text="
        + msg
        + "&app_absent=0"
    )
    print(url)

    driver.get(url)
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    time.sleep(5)

    try:
        send_key = wait.until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "/html/body/div[1]/div/div/div[5]/div/footer/div[1]/div/span[2]/div/div[2]/div[2]/button/span",
                )
            )
        )
        send_key.click()
        print("پیام با موفقیت به " + recipient + " ارسال شد!")
        time.sleep(2)
    except Exception as e:
        print("خطا در ارسال پیام به " + recipient)
        print("Error message:", str(e))


class Signal(QObject):
    list_updated = pyqtSignal()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("WhatsApp Message Sender")
        self.setup_ui()

        self.driver = None
        self.wait = None
        self.recipients = []
        self.message = ""
        self.list_signal = Signal()
        self.list_signal.list_updated.connect(self.update_list_item)

        self.setStyleSheet(dark_stylesheet)

    def setup_ui(self):
        main_widget = QWidget()

        self.file_entry = QLineEdit()
        self.file_entry.setPlaceholderText("Enter file path")
        self.file_entry.setStyleSheet(
            "QLineEdit::placeholder { color: #888888; font-family: Roboto, sans-serif;  }"
        )
        browse_button = QPushButton("Browse")
        browse_button.clicked.connect(self.browse_file)

        self.message_entry = QLineEdit()
        self.message_entry.setPlaceholderText("Enter your message")
        self.message_entry.setStyleSheet(
            "QLineEdit::placeholder { color: #888888; font-family: Roboto, sans-serif; font-weight: bold; }"
        )

        self.message_entry.setMinimumHeight(80)

        self.message_entry.setAlignment(Qt.AlignTop | Qt.AlignLeft)

        text = self.message_entry.text()
        if any(ord(c) > 127 for c in text):

            self.message_entry.setAlignment(Qt.AlignTop | Qt.AlignRight)

        start_sending_button = QPushButton("Start Sending")
        start_sending_button.clicked.connect(self.start_sending)

        self.code_scanned_button = QPushButton("Code Scanned")
        self.code_scanned_button.setEnabled(False)
        self.code_scanned_button.clicked.connect(self.code_scanned)

        self.sent_list = QListWidget()

        layout = QVBoxLayout()
        # layout.addWidget(file_label)
        layout.addWidget(self.file_entry)
        layout.addWidget(browse_button)
        # layout.addWidget(message_label)
        layout.addWidget(self.message_entry)
        layout.addWidget(start_sending_button)
        layout.addWidget(self.code_scanned_button)
        layout.addWidget(self.sent_list)
        label = QLabel()
        font = QtGui.QFont("Roboto", 22)
        label.setFont(font)
        main_widget.setLayout(layout)
        self.setCentralWidget(main_widget)

        self.setStyleSheet(dark_stylesheet)

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx)"
        )
        if file_path:
            self.file_entry.setText(file_path)

    def start_sending(self):
        self.recipients = self.read_contacts(self.file_entry.text())
        self.message = self.message_entry.text()

        self.driver = initialize_driver()
        self.wait = WebDriverWait(self.driver, 10)

        self.driver.get("https://web.whatsapp.com")
        time.sleep(2)

        self.code_scanned_button.setEnabled(True)

    def code_scanned(self):

        self.sending_thread = MessageSendingThread(
            self.driver, self.wait, self.recipients, self.message
        )
        self.sending_thread.update_list.connect(self.update_list_item)
        self.sending_thread.start()

    def read_contacts(self, file_path):
        values = []
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                value = row[0]
                values.append(str(value))

            workbook.close()

        except FileNotFoundError:
            self.show_message_box("Error", "File not found: " + file_path)
        except Exception as e:
            self.show_message_box(
                "Error", "An error occurred while reading the file: " + str(e)
            )

        return values

    def update_list_item(self, item):
        list_item = QListWidgetItem(str(item))
        self.sent_list.addItem(list_item)

    def show_message_box(self, title, message):
        messagebox = QMessageBox(self)
        messagebox.setWindowTitle(title)
        messagebox.setText(message)
        messagebox.setIcon(QMessageBox.Information)
        messagebox.exec_()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
